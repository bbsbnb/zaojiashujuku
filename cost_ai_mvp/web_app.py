from __future__ import annotations

import os
import shutil
import sys
import threading
import webbrowser
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.serving import make_server

from .cli import import_samples, price_file, query
from .db import connect, init_library
from .shaanxi_import import import_shaanxi_price
from .utils import new_id, now, safe_filename, today_compact


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LIBRARY = ROOT / "AI造价数据库"
DEFAULT_SAMPLES = ROOT / "第一轮测试样本"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024


@app.before_request
def _log_request() -> None:
    print(f"[web_app] request {request.method} {request.path}", flush=True)


@app.errorhandler(404)
def _log_404(error):
    print(f"[web_app] 404 {request.method} {request.path}", flush=True)
    return error


def _library() -> Path:
    return _resolve_library(request.values.get("library") or str(DEFAULT_LIBRARY))


def _resolve_library(value: str | Path) -> Path:
    path = Path(value)
    if not path.is_absolute():
        path = ROOT / path
    if "?" in str(path) and DEFAULT_LIBRARY.exists():
        return DEFAULT_LIBRARY
    return path


def _stats(library: Path) -> dict[str, int | str]:
    if not (library / "database" / "cost_database.sqlite").exists():
        return {"projects": 0, "files": 0, "prices": 0, "queries": 0, "tasks": 0, "library": str(library)}
    with connect(library) as conn:
        return {
            "projects": conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0],
            "files": conn.execute("SELECT COUNT(*) FROM files").fetchone()[0],
            "prices": conn.execute("SELECT COUNT(*) FROM prices").fetchone()[0],
            "queries": conn.execute("SELECT COUNT(*) FROM query_logs").fetchone()[0],
            "tasks": conn.execute("SELECT COUNT(*) FROM pricing_tasks").fetchone()[0],
            "library": str(library),
        }


def _price_data_quality(library: Path) -> dict[str, int]:
    if not (library / "database" / "cost_database.sqlite").exists():
        return {"missing_required": 0, "invalid_price": 0, "duplicate_items": 0, "outliers": 0}
    with connect(library) as conn:
        duplicate_items = conn.execute(
            """
            SELECT COUNT(*) FROM (
                SELECT item_type, COALESCE(standard_name, original_name), COALESCE(standard_spec, original_spec, ''),
                       COALESCE(standard_unit, original_unit), region, unit_price
                FROM prices
                GROUP BY item_type, COALESCE(standard_name, original_name), COALESCE(standard_spec, original_spec, ''),
                         COALESCE(standard_unit, original_unit), region, unit_price
                HAVING COUNT(*) > 1
            )
            """
        ).fetchone()[0]
        return {
            "missing_required": conn.execute(
                """
                SELECT COUNT(*) FROM prices
                WHERE COALESCE(original_name, standard_name, '') = ''
                   OR COALESCE(original_unit, standard_unit, '') = ''
                   OR COALESCE(region, '') = ''
                   OR COALESCE(item_type, '') = ''
                """
            ).fetchone()[0],
            "invalid_price": conn.execute("SELECT COUNT(*) FROM prices WHERE unit_price IS NULL OR unit_price <= 0").fetchone()[0],
            "duplicate_items": duplicate_items,
            "outliers": conn.execute("SELECT COUNT(*) FROM prices WHERE is_outlier = 1").fetchone()[0],
        }


def _price_filter_query(filters: dict[str, str | int], limit: int | None = None) -> tuple[str, list[object]]:
    where = ["1=1"]
    params: list[object] = []
    if filters.get("q"):
        where.append("(p.original_name LIKE ? OR p.standard_name LIKE ? OR p.original_spec LIKE ? OR p.standard_spec LIKE ?)")
        like = f"%{filters['q']}%"
        params.extend([like, like, like, like])
    if filters.get("item_type"):
        where.append("p.item_type = ?")
        params.append(filters["item_type"])
    if filters.get("region"):
        where.append("p.region = ?")
        params.append(filters["region"])
    if filters.get("source_type"):
        where.append("p.source_type = ?")
        params.append(filters["source_type"])
    sql = f"""
        SELECT p.*, f.archive_path, f.original_name AS file_name
        FROM prices p
        LEFT JOIN files f ON f.id = p.file_id
        WHERE {' AND '.join(where)}
        ORDER BY p.created_at DESC
    """
    if limit is not None:
        sql += " LIMIT ?"
        params.append(limit)
    return sql, params


def _ensure_p2_schema(library: Path) -> None:
    init_library(library)
    with connect(library) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS material_market_prices (
                id TEXT PRIMARY KEY,
                item_name TEXT NOT NULL,
                spec TEXT,
                unit TEXT NOT NULL,
                region TEXT NOT NULL,
                market_date TEXT NOT NULL,
                market_price REAL NOT NULL,
                source_name TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS regional_adjustments (
                id TEXT PRIMARY KEY,
                region TEXT NOT NULL,
                adjustment_name TEXT NOT NULL,
                adjustment_type TEXT NOT NULL,
                coefficient REAL NOT NULL,
                effective_date TEXT NOT NULL,
                policy_no TEXT,
                notes TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_market_filter ON material_market_prices(item_name, region, market_date);
            CREATE INDEX IF NOT EXISTS idx_adjustments_region ON regional_adjustments(region, is_active, effective_date);
            """
        )
        conn.commit()


def _ensure_p3_schema(library: Path) -> None:
    _ensure_p2_schema(library)
    with connect(library) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS audit_logs (
                id TEXT PRIMARY KEY,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id TEXT,
                summary TEXT NOT NULL,
                user_name TEXT DEFAULT 'local_user',
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_audit_logs_created ON audit_logs(created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_audit_logs_entity ON audit_logs(entity_type, entity_id);
            """
        )
        conn.commit()


def _write_audit_log(library: Path, action: str, entity_type: str, entity_id: str | None, summary: str) -> None:
    _ensure_p3_schema(library)
    with connect(library) as conn:
        conn.execute(
            """
            INSERT INTO audit_logs (id, action, entity_type, entity_id, summary, user_name, created_at)
            VALUES (?, ?, ?, ?, ?, 'local_user', ?)
            """,
            (new_id("AUD"), action, entity_type, entity_id, summary, now()),
        )
        conn.commit()


def _recommend_prices(library: Path, text: str, region: str = "", item_type: str = "") -> list[dict]:
    if not text.strip() or not (library / "database" / "cost_database.sqlite").exists():
        return []
    where = ["p.is_outlier = 0", "(p.original_name LIKE ? OR p.standard_name LIKE ? OR p.original_spec LIKE ? OR p.standard_spec LIKE ?)"]
    like = f"%{text.strip()}%"
    params: list[object] = [like, like, like, like]
    if region:
        where.append("p.region = ?")
        params.append(region)
    if item_type:
        where.append("p.item_type = ?")
        params.append(item_type)
    with connect(library) as conn:
        rows = [dict(r) for r in conn.execute(
            f"""
            SELECT
                p.item_type,
                COALESCE(p.standard_name, p.original_name) AS item_name,
                COALESCE(p.standard_spec, p.original_spec, '') AS spec,
                COALESCE(p.standard_unit, p.original_unit) AS unit,
                p.region,
                COUNT(*) AS sample_count,
                MIN(p.unit_price) AS low_price,
                AVG(p.unit_price) AS avg_price,
                MAX(p.unit_price) AS high_price,
                MAX(COALESCE(p.price_date, CAST(p.price_year AS TEXT), p.created_at)) AS latest_time
            FROM prices p
            WHERE {' AND '.join(where)}
            GROUP BY p.item_type, COALESCE(p.standard_name, p.original_name), COALESCE(p.standard_spec, p.original_spec, ''), COALESCE(p.standard_unit, p.original_unit), p.region
            ORDER BY sample_count DESC, latest_time DESC
            LIMIT 20
            """,
            params,
        ).fetchall()]
    for row in rows:
        count = row.get("sample_count") or 0
        spread = (row.get("high_price") or 0) - (row.get("low_price") or 0)
        avg_price = row.get("avg_price") or 0
        if count >= 5 and (not avg_price or spread / avg_price <= 0.3):
            row["confidence"] = "高"
        elif count >= 3:
            row["confidence"] = "中"
        else:
            row["confidence"] = "低"
        row["recommend_reason"] = f"匹配历史样本 {count} 条，参考区间 {row.get('low_price') or 0:.2f}-{row.get('high_price') or 0:.2f}。"
    return rows


def _pricing_task_summary(library: Path, task_id: str) -> dict | None:
    if not (library / "database" / "cost_database.sqlite").exists():
        return None
    with connect(library) as conn:
        task = conn.execute("SELECT * FROM pricing_tasks WHERE id = ?", (task_id,)).fetchone()
        if not task:
            return None
        task = dict(task)
        item_rows = [dict(r) for r in conn.execute(
            """
            SELECT
                i.*,
                i.is_accepted,
                COALESCE(COUNT(c.id), 0) AS candidate_count
            FROM pricing_task_items i
            LEFT JOIN match_candidates c ON c.task_item_id = i.id
            WHERE i.task_id = ?
            GROUP BY i.id
            ORDER BY i.source_row
            """,
            (task_id,),
        ).fetchall()]
        candidate_rows = [dict(r) for r in conn.execute(
            """
            SELECT c.*
            FROM match_candidates c
            JOIN pricing_task_items i ON i.id = c.task_item_id
            WHERE i.task_id = ?
            ORDER BY i.source_row, c.rank_no
            """,
            (task_id,),
        ).fetchall()]
    candidates_by_item: dict[str, list[dict]] = {}
    for row in candidate_rows:
        candidates_by_item.setdefault(row["task_item_id"], []).append(row)
    grouped_items: dict[str, list[dict]] = {"labor": [], "material": [], "machinery": [], "other": []}
    for row in item_rows:
        key = row.get("item_type") or "other"
        if key not in grouped_items:
            key = "other"
        grouped_items[key].append(row)
    return {"task": task, "items": item_rows, "grouped_items": grouped_items, "candidates_by_item": candidates_by_item}


def _task_rows(library: Path) -> list[dict]:
    if not (library / "database" / "cost_database.sqlite").exists():
        return []
    with connect(library) as conn:
        return [dict(r) for r in conn.execute(
            """
            SELECT
                t.*,
                COALESCE(i.item_count, 0) AS item_count,
                COALESCE(i.pending_count, 0) AS pending_count,
                COALESCE(i.confirmed_count, 0) AS confirmed_count,
                COALESCE(c.candidate_count, 0) AS candidate_count
            FROM pricing_tasks t
            LEFT JOIN (
                SELECT
                    task_id,
                    COUNT(*) AS item_count,
                    SUM(CASE WHEN status = 'confirmed' OR is_accepted = 1 THEN 0 ELSE 1 END) AS pending_count,
                    SUM(CASE WHEN status = 'confirmed' OR is_accepted = 1 THEN 1 ELSE 0 END) AS confirmed_count
                FROM pricing_task_items
                GROUP BY task_id
            ) i ON i.task_id = t.id
            LEFT JOIN (
                SELECT i.task_id, COUNT(c.id) AS candidate_count
                FROM pricing_task_items i
                LEFT JOIN match_candidates c ON c.task_item_id = i.id
                GROUP BY i.task_id
            ) c ON c.task_id = t.id
            ORDER BY t.created_at DESC
            """
        ).fetchall()]


def _settings(library: Path) -> dict[str, str]:
    if not (library / "database" / "cost_database.sqlite").exists():
        return {}
    with connect(library) as conn:
        return {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM app_settings").fetchall()}


@app.get("/")
def index():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    recent_queries = []
    recent_tasks = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            recent_queries = [dict(r) for r in conn.execute(
                "SELECT query_text, suggested_price, reference_low, reference_high, confidence, created_at FROM query_logs ORDER BY created_at DESC LIMIT 5"
            ).fetchall()]
            recent_tasks = [dict(r) for r in conn.execute(
                "SELECT id, project_name, region, project_type, status, created_at FROM pricing_tasks ORDER BY created_at DESC LIMIT 5"
            ).fetchall()]
    return render_template(
        "index.html",
        stats=_stats(library),
        default_samples=str(DEFAULT_SAMPLES),
        recent_queries=recent_queries,
        recent_tasks=recent_tasks,
    )


@app.route("/projects", methods=["GET"])
def projects_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    filters = {
        "q": request.args.get("q", "").strip(),
        "region": request.args.get("region", "").strip(),
        "year": request.args.get("year", "").strip(),
        "project_type": request.args.get("project_type", "").strip(),
    }
    rows = []
    regions: list[str] = []
    years: list[int] = []
    project_types: list[str] = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            regions = [r[0] for r in conn.execute("SELECT DISTINCT region FROM projects WHERE region IS NOT NULL ORDER BY region").fetchall()]
            years = [r[0] for r in conn.execute("SELECT DISTINCT year FROM projects ORDER BY year DESC").fetchall()]
            project_types = [r[0] for r in conn.execute("SELECT DISTINCT project_type FROM projects WHERE project_type IS NOT NULL ORDER BY project_type").fetchall()]
            where = ["1=1"]
            params: list[object] = []
            if filters["q"]:
                where.append("(p.name LIKE ? OR p.short_name LIKE ?)")
                like = f"%{filters['q']}%"
                params.extend([like, like])
            if filters["region"]:
                where.append("p.region = ?")
                params.append(filters["region"])
            if filters["year"]:
                where.append("p.year = ?")
                params.append(int(filters["year"]))
            if filters["project_type"]:
                where.append("p.project_type = ?")
                params.append(filters["project_type"])
            rows = [dict(r) for r in conn.execute(
                f"""
                SELECT
                    p.*,
                    COALESCE(fstats.file_count, 0) AS file_count,
                    COALESCE(pstats.price_count, 0) AS price_count,
                    COALESCE(fstats.latest_imported_at, '') AS latest_imported_at
                FROM projects p
                LEFT JOIN (
                    SELECT project_id, COUNT(*) AS file_count, MAX(imported_at) AS latest_imported_at
                    FROM files
                    GROUP BY project_id
                ) fstats ON fstats.project_id = p.id
                LEFT JOIN (
                    SELECT project_id, COUNT(*) AS price_count
                    FROM prices
                    GROUP BY project_id
                ) pstats ON pstats.project_id = p.id
                WHERE {' AND '.join(where)}
                ORDER BY p.updated_at DESC, p.created_at DESC
                """,
                params,
            ).fetchall()]
    return render_template(
        "projects.html",
        rows=rows,
        filters=filters,
        regions=regions,
        years=years,
        project_types=project_types,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/projects/<project_id>", methods=["GET"])
def project_detail_route(project_id: str):
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    project = None
    files = []
    prices = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if project:
                project = dict(project)
                files = [dict(r) for r in conn.execute("SELECT * FROM files WHERE project_id = ? ORDER BY created_at DESC", (project_id,)).fetchall()]
                prices = [dict(r) for r in conn.execute("SELECT * FROM prices WHERE project_id = ? AND is_outlier = 0 ORDER BY created_at DESC LIMIT 200", (project_id,)).fetchall()]
    return render_template(
        "project_detail.html",
        project=project,
        files=files,
        prices=prices,
        stats=_stats(library),
        back=url_for("projects_route", library=str(library)),
    )


@app.route("/analytics", methods=["GET"])
def analytics_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    filters = {
        "q": request.args.get("q", "").strip(),
        "region": request.args.get("region", "").strip(),
        "project_type": request.args.get("project_type", "").strip(),
        "item_type": request.args.get("item_type", "").strip(),
    }
    project_rows = []
    trend_rows = []
    indicator_rows = []
    regions: list[str] = []
    project_types: list[str] = []
    item_types: list[str] = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            regions = [r[0] for r in conn.execute("SELECT DISTINCT region FROM prices WHERE region IS NOT NULL ORDER BY region").fetchall()]
            project_types = [r[0] for r in conn.execute("SELECT DISTINCT project_type FROM projects WHERE project_type IS NOT NULL ORDER BY project_type").fetchall()]
            item_types = [r[0] for r in conn.execute("SELECT DISTINCT item_type FROM prices WHERE item_type IS NOT NULL ORDER BY item_type").fetchall()]
            project_where = ["p.status = 'active'"]
            project_params: list[object] = []
            if filters["region"]:
                project_where.append("p.region = ?")
                project_params.append(filters["region"])
            if filters["project_type"]:
                project_where.append("p.project_type = ?")
                project_params.append(filters["project_type"])
            project_rows = [dict(r) for r in conn.execute(
                f"""
                SELECT
                    p.id,
                    p.name,
                    p.region,
                    p.year,
                    p.project_type,
                    p.building_area,
                    COUNT(pr.id) AS price_count,
                    SUM(COALESCE(pr.total_price, pr.quantity * pr.unit_price, 0)) AS total_cost,
                    CASE WHEN p.building_area > 0
                        THEN SUM(COALESCE(pr.total_price, pr.quantity * pr.unit_price, 0)) / p.building_area
                        ELSE NULL
                    END AS cost_per_area
                FROM projects p
                LEFT JOIN prices pr ON pr.project_id = p.id AND pr.is_outlier = 0
                WHERE {' AND '.join(project_where)}
                GROUP BY p.id
                ORDER BY cost_per_area DESC, p.updated_at DESC
                LIMIT 80
                """,
                project_params,
            ).fetchall()]
            trend_where = ["p.is_outlier = 0"]
            trend_params: list[object] = []
            if filters["q"]:
                trend_where.append("(p.original_name LIKE ? OR p.standard_name LIKE ? OR p.original_spec LIKE ? OR p.standard_spec LIKE ?)")
                like = f"%{filters['q']}%"
                trend_params.extend([like, like, like, like])
            if filters["region"]:
                trend_where.append("p.region = ?")
                trend_params.append(filters["region"])
            if filters["item_type"]:
                trend_where.append("p.item_type = ?")
                trend_params.append(filters["item_type"])
            trend_rows = [dict(r) for r in conn.execute(
                f"""
                SELECT
                    COALESCE(CAST(p.price_year AS TEXT), substr(p.price_date, 1, 4), '未知') AS period,
                    COUNT(*) AS sample_count,
                    MIN(p.unit_price) AS min_price,
                    AVG(p.unit_price) AS avg_price,
                    MAX(p.unit_price) AS max_price
                FROM prices p
                WHERE {' AND '.join(trend_where)}
                GROUP BY COALESCE(CAST(p.price_year AS TEXT), substr(p.price_date, 1, 4), '未知')
                ORDER BY period
                """,
                trend_params,
            ).fetchall()]
            indicator_rows = [dict(r) for r in conn.execute(
                """
                SELECT
                    p.project_type,
                    COUNT(DISTINCT p.id) AS project_count,
                    COUNT(pr.id) AS price_count,
                    AVG(pr.unit_price) AS avg_unit_price,
                    AVG(CASE WHEN p.building_area > 0 THEN COALESCE(pr.total_price, pr.quantity * pr.unit_price, 0) / p.building_area END) AS avg_cost_per_area
                FROM projects p
                JOIN prices pr ON pr.project_id = p.id AND pr.is_outlier = 0
                GROUP BY p.project_type
                ORDER BY project_count DESC, price_count DESC
                LIMIT 30
                """
            ).fetchall()]
    return render_template(
        "analytics.html",
        filters=filters,
        project_rows=project_rows,
        trend_rows=trend_rows,
        indicator_rows=indicator_rows,
        regions=regions,
        project_types=project_types,
        item_types=item_types,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/standard-items", methods=["GET", "POST"])
def standard_items_route():
    library = _library() if request.method == "POST" else _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    init_library(library)
    errors: list[str] = []
    if request.method == "POST":
        item_type = request.form.get("item_type", "").strip()
        standard_name = request.form.get("standard_name", "").strip()
        standard_unit = request.form.get("standard_unit", "").strip()
        standard_spec = request.form.get("standard_spec", "").strip()
        if not item_type:
            errors.append("类型不能为空")
        if not standard_name:
            errors.append("标准名称不能为空")
        if not standard_unit:
            errors.append("单位不能为空")
        if not errors:
            with connect(library) as conn:
                exists = conn.execute(
                    """
                    SELECT id FROM standard_items
                    WHERE item_type = ? AND standard_name = ? AND COALESCE(standard_spec, '') = ? AND standard_unit = ? AND is_active = 1
                    """,
                    (item_type, standard_name, standard_spec, standard_unit),
                ).fetchone()
                if exists:
                    errors.append("相同类型、名称、规格、单位的标准项已存在")
                else:
                    current = now()
                    conn.execute(
                        """
                        INSERT INTO standard_items (
                            id, item_type, standard_name, standard_spec, standard_unit, category,
                            keywords, aliases, brand_requirement, notes, created_by, is_active, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (
                            new_id("STD"),
                            item_type,
                            standard_name,
                            standard_spec,
                            standard_unit,
                            request.form.get("category", "").strip(),
                            request.form.get("keywords", "").strip(),
                            request.form.get("aliases", "").strip(),
                            request.form.get("brand_requirement", "").strip(),
                            request.form.get("notes", "").strip(),
                            "web",
                            current,
                            current,
                        ),
                    )
                    conn.commit()
                    _write_audit_log(library, "create", "standard_item", standard_name, f"新增标准项：{standard_name}")
                    return redirect(url_for("standard_items_route", library=str(library)))
    rows = []
    inferred_rows = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            rows = [dict(r) for r in conn.execute("SELECT * FROM standard_items WHERE is_active = 1 ORDER BY updated_at DESC LIMIT 100").fetchall()]
            inferred_rows = [dict(r) for r in conn.execute(
                """
                SELECT
                    item_type,
                    COALESCE(standard_name, original_name) AS standard_name,
                    COALESCE(standard_spec, original_spec, '') AS standard_spec,
                    COALESCE(standard_unit, original_unit) AS standard_unit,
                    COUNT(*) AS source_count,
                    AVG(unit_price) AS avg_price
                FROM prices
                WHERE is_outlier = 0
                GROUP BY item_type, COALESCE(standard_name, original_name), COALESCE(standard_spec, original_spec, ''), COALESCE(standard_unit, original_unit)
                ORDER BY source_count DESC
                LIMIT 30
                """
            ).fetchall()]
    return render_template(
        "standard_items.html",
        rows=rows,
        inferred_rows=inferred_rows,
        errors=errors,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/market", methods=["GET", "POST"])
def market_route():
    library = _library() if request.method == "POST" else _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p2_schema(library)
    errors: list[str] = []
    if request.method == "POST":
        item_name = request.form.get("item_name", "").strip()
        unit = request.form.get("unit", "").strip()
        region = request.form.get("region", "").strip()
        market_date = request.form.get("market_date", "").strip()
        price_text = request.form.get("market_price", "").strip()
        if not item_name:
            errors.append("材料名称不能为空")
        if not unit:
            errors.append("单位不能为空")
        if not region:
            errors.append("地区不能为空")
        if not market_date:
            errors.append("行情日期不能为空")
        try:
            market_price = float(price_text)
            if market_price <= 0:
                errors.append("市场价必须大于0")
        except ValueError:
            market_price = 0.0
            errors.append("市场价必须是数字")
        if not errors:
            current = now()
            with connect(library) as conn:
                conn.execute(
                    """
                    INSERT INTO material_market_prices (
                        id, item_name, spec, unit, region, market_date, market_price,
                        source_name, notes, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_id("MKT"),
                        item_name,
                        request.form.get("spec", "").strip(),
                        unit,
                        region,
                        market_date,
                        market_price,
                        request.form.get("source_name", "").strip(),
                        request.form.get("notes", "").strip(),
                        current,
                        current,
                    ),
                )
                conn.commit()
            _write_audit_log(library, "create", "material_market_price", item_name, f"新增材料行情：{item_name} {market_price}")
            return redirect(url_for("market_route", library=str(library)))
    rows = []
    compare_rows = []
    trend_rows = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            rows = [dict(r) for r in conn.execute("SELECT * FROM material_market_prices ORDER BY market_date DESC, updated_at DESC LIMIT 80").fetchall()]
            compare_rows = [dict(r) for r in conn.execute(
                """
                SELECT
                    m.item_name,
                    m.spec,
                    m.unit,
                    m.region,
                    m.market_date,
                    m.market_price,
                    AVG(p.unit_price) AS history_avg_price,
                    COUNT(p.id) AS history_count,
                    CASE WHEN AVG(p.unit_price) > 0 THEN (m.market_price - AVG(p.unit_price)) / AVG(p.unit_price) * 100 END AS diff_percent
                FROM material_market_prices m
                LEFT JOIN prices p ON p.is_outlier = 0
                    AND p.region = m.region
                    AND COALESCE(p.standard_unit, p.original_unit) = m.unit
                    AND (p.standard_name LIKE '%' || m.item_name || '%' OR p.original_name LIKE '%' || m.item_name || '%')
                GROUP BY m.id
                ORDER BY m.market_date DESC
                LIMIT 80
                """
            ).fetchall()]
            trend_rows = [dict(r) for r in conn.execute(
                """
                SELECT item_name, region, market_date, AVG(market_price) AS avg_price, COUNT(*) AS sample_count
                FROM material_market_prices
                GROUP BY item_name, region, market_date
                ORDER BY market_date DESC
                LIMIT 80
                """
            ).fetchall()]
    return render_template(
        "market.html",
        rows=rows,
        compare_rows=compare_rows,
        trend_rows=trend_rows,
        errors=errors,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/adjustments", methods=["GET", "POST"])
def adjustments_route():
    library = _library() if request.method == "POST" else _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p2_schema(library)
    errors: list[str] = []
    if request.method == "POST":
        region = request.form.get("region", "").strip()
        adjustment_name = request.form.get("adjustment_name", "").strip()
        adjustment_type = request.form.get("adjustment_type", "").strip()
        effective_date = request.form.get("effective_date", "").strip()
        coefficient_text = request.form.get("coefficient", "").strip()
        if not region:
            errors.append("地区不能为空")
        if not adjustment_name:
            errors.append("调价名称不能为空")
        if adjustment_type not in {"region", "labor", "material", "machinery", "tax"}:
            errors.append("调价类型无效")
        if not effective_date:
            errors.append("生效日期不能为空")
        try:
            coefficient = float(coefficient_text)
            if coefficient <= 0:
                errors.append("系数必须大于0")
        except ValueError:
            coefficient = 1.0
            errors.append("系数必须是数字")
        if not errors:
            current = now()
            with connect(library) as conn:
                conn.execute(
                    """
                    INSERT INTO regional_adjustments (
                        id, region, adjustment_name, adjustment_type, coefficient,
                        effective_date, policy_no, notes, is_active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        new_id("ADJ"),
                        region,
                        adjustment_name,
                        adjustment_type,
                        coefficient,
                        effective_date,
                        request.form.get("policy_no", "").strip(),
                        request.form.get("notes", "").strip(),
                        current,
                        current,
                    ),
                )
                conn.commit()
            _write_audit_log(library, "create", "regional_adjustment", adjustment_name, f"新增调价规则：{adjustment_name} 系数 {coefficient}")
            return redirect(url_for("adjustments_route", library=str(library)))
    rows = []
    preview_rows = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            rows = [dict(r) for r in conn.execute("SELECT * FROM regional_adjustments WHERE is_active = 1 ORDER BY effective_date DESC, updated_at DESC LIMIT 80").fetchall()]
            preview_rows = [dict(r) for r in conn.execute(
                """
                SELECT
                    p.region,
                    p.item_type,
                    COALESCE(p.standard_name, p.original_name) AS item_name,
                    COALESCE(p.standard_unit, p.original_unit) AS unit,
                    p.unit_price,
                    a.adjustment_name,
                    a.coefficient,
                    ROUND(p.unit_price * a.coefficient, 2) AS adjusted_price
                FROM prices p
                JOIN regional_adjustments a ON a.region = p.region
                    AND a.is_active = 1
                    AND (a.adjustment_type = 'region' OR a.adjustment_type = p.item_type)
                WHERE p.is_outlier = 0
                ORDER BY a.effective_date DESC, p.created_at DESC
                LIMIT 80
                """
            ).fetchall()]
    return render_template(
        "adjustments.html",
        rows=rows,
        preview_rows=preview_rows,
        errors=errors,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/reports", methods=["GET"])
def reports_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p2_schema(library)
    return render_template("reports.html", stats=_stats(library), back=url_for("index", library=str(library)))


@app.route("/export/report", methods=["GET"])
def export_report_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p2_schema(library)
    report_type = request.args.get("type", "summary")
    wb = Workbook()
    header_fill = PatternFill(start_color="00957F", end_color="00957F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def write_table(ws, headers: list[str], rows: list[list[object]]) -> None:
        for col_index, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(rows, 2):
            for col_index, value in enumerate(row, 1):
                ws.cell(row=row_index, column=col_index, value=value)
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)

    with connect(library) as conn:
        summary_stats = _stats(library)
        cover = wb.active
        cover.title = "报表摘要"
        write_table(
            cover,
            ["项目", "数值"],
            [
                ["报表类型", "造价数据库综合报表" if report_type == "summary" else report_type],
                ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["项目数", summary_stats["projects"]],
                ["文件数", summary_stats["files"]],
                ["价格记录", summary_stats["prices"]],
                ["资料库", summary_stats["library"]],
            ],
        )

        prices_ws = wb.create_sheet("价格库")
        price_rows = [
            [
                r["item_type"],
                r["standard_name"] or r["original_name"],
                r["standard_spec"] or r["original_spec"],
                r["standard_unit"] or r["original_unit"],
                r["unit_price"],
                r["region"],
                r["price_date"] or r["price_year"],
                r["source_type"],
            ]
            for r in conn.execute("SELECT * FROM prices WHERE is_outlier = 0 ORDER BY created_at DESC LIMIT 1000").fetchall()
        ]
        write_table(prices_ws, ["类型", "名称", "规格", "单位", "单价", "地区", "年份/日期", "来源"], price_rows)

        project_ws = wb.create_sheet("项目指标")
        project_rows = [
            [r["name"], r["region"], r["year"], r["project_type"], r["building_area"], r["price_count"], round(r["total_cost"] or 0, 2), round(r["cost_per_area"] or 0, 2)]
            for r in conn.execute(
                """
                SELECT p.name, p.region, p.year, p.project_type, p.building_area,
                       COUNT(pr.id) AS price_count,
                       SUM(COALESCE(pr.total_price, pr.quantity * pr.unit_price, 0)) AS total_cost,
                       CASE WHEN p.building_area > 0 THEN SUM(COALESCE(pr.total_price, pr.quantity * pr.unit_price, 0)) / p.building_area END AS cost_per_area
                FROM projects p
                LEFT JOIN prices pr ON pr.project_id = p.id AND pr.is_outlier = 0
                GROUP BY p.id
                ORDER BY p.updated_at DESC
                LIMIT 200
                """
            ).fetchall()
        ]
        write_table(project_ws, ["项目名称", "地区", "年份", "类型", "面积", "价格数", "总价合计", "单方指标"], project_rows)

        market_ws = wb.create_sheet("材料行情")
        market_rows = [
            [r["item_name"], r["spec"], r["unit"], r["region"], r["market_date"], r["market_price"], r["source_name"], r["notes"]]
            for r in conn.execute("SELECT * FROM material_market_prices ORDER BY market_date DESC LIMIT 1000").fetchall()
        ]
        write_table(market_ws, ["材料名称", "规格", "单位", "地区", "行情日期", "市场价", "来源", "备注"], market_rows)

        adjustment_ws = wb.create_sheet("调价政策")
        adjustment_rows = [
            [r["region"], r["adjustment_name"], r["adjustment_type"], r["coefficient"], r["effective_date"], r["policy_no"], r["notes"]]
            for r in conn.execute("SELECT * FROM regional_adjustments WHERE is_active = 1 ORDER BY effective_date DESC LIMIT 500").fetchall()
        ]
        write_table(adjustment_ws, ["地区", "调价名称", "类型", "系数", "生效日期", "政策编号", "备注"], adjustment_rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"造价数据库综合报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/smart-pricing", methods=["GET", "POST"])
def smart_pricing_route():
    library = _library() if request.method == "POST" else _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    text = request.values.get("text", "").strip()
    region = request.values.get("region", "").strip()
    item_type = request.values.get("item_type", "").strip()
    recommendations = _recommend_prices(library, text, region, item_type) if text else []
    if request.method == "POST" and text:
        _write_audit_log(library, "recommend", "smart_pricing", text, f"智能组价推荐：{text}，返回 {len(recommendations)} 条")
    regions = []
    item_types = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            regions = [r[0] for r in conn.execute("SELECT DISTINCT region FROM prices WHERE region IS NOT NULL ORDER BY region").fetchall()]
            item_types = [r[0] for r in conn.execute("SELECT DISTINCT item_type FROM prices WHERE item_type IS NOT NULL ORDER BY item_type").fetchall()]
    return render_template(
        "smart_pricing.html",
        text=text,
        region=region,
        item_type=item_type,
        recommendations=recommendations,
        regions=regions,
        item_types=item_types,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/audit", methods=["GET"])
def audit_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    rows = []
    with connect(library) as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 200").fetchall()]
    return render_template("audit.html", rows=rows, stats=_stats(library), back=url_for("index", library=str(library)))


@app.route("/api-docs", methods=["GET"])
def api_docs_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    return render_template("api_docs.html", stats=_stats(library), back=url_for("index", library=str(library)))


@app.route("/api/stats", methods=["GET"])
def api_stats_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    return jsonify({"code": 0, "data": _stats(library)})


@app.route("/api/prices", methods=["GET"])
def api_prices_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    filters = {
        "q": request.args.get("q", "").strip(),
        "item_type": request.args.get("item_type", ""),
        "region": request.args.get("region", ""),
        "source_type": request.args.get("source_type", ""),
    }
    limit = min(int(request.args.get("limit", "100") or 100), 500)
    rows = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            sql, params = _price_filter_query(filters, limit)
            rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    return jsonify({"code": 0, "data": rows, "limit": limit})


@app.route("/api/recommend", methods=["GET"])
def api_recommend_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    _ensure_p3_schema(library)
    text = request.args.get("q", "").strip()
    region = request.args.get("region", "").strip()
    item_type = request.args.get("item_type", "").strip()
    rows = _recommend_prices(library, text, region, item_type)
    if text:
        _write_audit_log(library, "api_recommend", "api", text, f"API推荐：{text}，返回 {len(rows)} 条")
    return jsonify({"code": 0, "data": rows})


@app.route("/tasks", methods=["GET"])
def tasks_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    status_filter = request.args.get("status", "").strip()
    rows = _task_rows(library)
    if status_filter == "pending":
        rows = [r for r in rows if (r.get("pending_count") or 0) > 0]
    elif status_filter == "confirmed":
        rows = [r for r in rows if (r.get("item_count") or 0) > 0 and (r.get("pending_count") or 0) == 0]
    return render_template(
        "tasks.html",
        rows=rows,
        status_filter=status_filter,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/match", methods=["GET"])
def match_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    rows = _task_rows(library)
    return render_template(
        "match.html",
        rows=rows,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/rules", methods=["GET"])
def rules_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    summary = {"active_items": 0, "alias_items": 0, "history_rules": 0, "price_names": 0}
    rows = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            summary["active_items"] = conn.execute("SELECT COUNT(*) FROM standard_items WHERE is_active = 1").fetchone()[0]
            summary["alias_items"] = conn.execute("SELECT COUNT(*) FROM standard_items WHERE aliases IS NOT NULL AND aliases != ''").fetchone()[0]
            summary["history_rules"] = conn.execute("SELECT COUNT(*) FROM name_match_history").fetchone()[0]
            summary["price_names"] = conn.execute("SELECT COUNT(DISTINCT COALESCE(standard_name, original_name)) FROM prices WHERE is_outlier = 0").fetchone()[0]
            rows = [dict(r) for r in conn.execute(
                """
                SELECT
                    item_type,
                    COALESCE(standard_name, original_name) AS standard_name,
                    COALESCE(standard_spec, original_spec, '') AS standard_spec,
                    COALESCE(standard_unit, original_unit) AS standard_unit,
                    COUNT(*) AS source_count,
                    MIN(unit_price) AS min_price,
                    AVG(unit_price) AS avg_price,
                    MAX(unit_price) AS max_price,
                    GROUP_CONCAT(DISTINCT region) AS regions
                FROM prices
                WHERE is_outlier = 0
                GROUP BY item_type, COALESCE(standard_name, original_name), COALESCE(standard_spec, original_spec, ''), COALESCE(standard_unit, original_unit)
                ORDER BY source_count DESC, standard_name
                LIMIT 80
                """
            ).fetchall()]
    return render_template(
        "rules.html",
        summary=summary,
        rows=rows,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/backup-ai", methods=["GET"])
def backup_ai_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    settings = _settings(library)
    db_path = library / "database" / "cost_database.sqlite"
    backup_dir = library / "backups" / "database"
    backups = []
    if backup_dir.exists():
        for path in sorted([p for p in backup_dir.iterdir() if p.is_file()], key=lambda p: p.stat().st_mtime, reverse=True)[:8]:
            stat = path.stat()
            backups.append({
                "name": path.name,
                "size_mb": stat.st_size / 1024 / 1024,
                "updated_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    return render_template(
        "backup_ai.html",
        settings=settings,
        db_path=db_path,
        db_size=db_path.stat().st_size if db_path.exists() else 0,
        backups=backups,
        backup_dir=backup_dir,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.post("/backup")
def backup_route():
    library = _library()
    init_library(library)
    db_path = library / "database" / "cost_database.sqlite"
    if not db_path.exists():
        return render_template("result.html", title="备份失败", message="资料库尚未初始化。", back=url_for("backup_ai_route", library=str(library)))
    backup_dir = library / "backups" / "database"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"cost_database_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sqlite"
    shutil.copy2(db_path, backup_path)
    return render_template("result.html", title="备份完成", message=f"已生成数据库备份：{backup_path}", back=url_for("backup_ai_route", library=str(library)))


@app.route("/tasks/<task_id>", methods=["GET"])
def task_detail_route(task_id: str):
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    summary = _pricing_task_summary(library, task_id)
    return render_template(
        "task_detail.html",
        summary=summary,
        stats=_stats(library),
        back=url_for("tasks_route", library=str(library)),
    )


@app.post("/tasks/<task_id>/select-candidate")
def select_candidate_route(task_id: str):
    library = _library()
    task_item_id = request.form.get("task_item_id", "").strip()
    candidate_id = request.form.get("candidate_id", "").strip()
    if not task_item_id or not candidate_id:
        return redirect(url_for("task_detail_route", task_id=task_id, library=str(library)))
    with connect(library) as conn:
        candidate = conn.execute(
            "SELECT * FROM match_candidates WHERE id = ? AND task_item_id = ?",
            (candidate_id, task_item_id),
        ).fetchone()
        if candidate:
            candidate = dict(candidate)
            conn.execute("UPDATE match_candidates SET is_selected = 0 WHERE task_item_id = ?", (task_item_id,))
            conn.execute("UPDATE match_candidates SET is_selected = 1 WHERE id = ?", (candidate_id,))
            conn.execute(
                """
                UPDATE pricing_task_items
                SET selected_standard_item_id = ?, standard_name = ?, standard_spec = ?, standard_unit = ?,
                    suggested_price = ?, reference_low = ?, reference_high = ?, source_type = ?, confidence = ?,
                    risk_tags = ?, risk_text = ?, is_accepted = 1, confirmed_price = ?, status = 'confirmed', updated_at = datetime('now')
                WHERE id = ?
                """,
                (
                    candidate.get("standard_item_id"),
                    candidate.get("candidate_name"),
                    candidate.get("candidate_spec"),
                    candidate.get("candidate_unit"),
                    candidate.get("candidate_price"),
                    candidate.get("reference_low"),
                    candidate.get("reference_high"),
                    candidate.get("source_type"),
                    candidate.get("confidence"),
                    candidate.get("risk_tags"),
                    candidate.get("risk_text"),
                    candidate.get("candidate_price"),
                    task_item_id,
                ),
            )
            conn.execute(
                "UPDATE pricing_tasks SET updated_at = datetime('now') WHERE id = ?",
                (task_id,),
            )
            conn.commit()
    return redirect(url_for("task_detail_route", task_id=task_id, library=str(library)))


@app.post("/init")
def init():
    library = _library()
    init_library(library)
    return render_template("result.html", title="资料库初始化", message=f"资料库已初始化：{library.resolve()}", back=url_for("index", library=str(library)))


@app.post("/import-samples")
def import_samples_route():
    library = _library()
    samples = Path(request.form.get("samples") or DEFAULT_SAMPLES)
    init_library(library)
    result = import_samples(library, samples)
    return render_template("import_result.html", result=result, stats=_stats(library), back=url_for("index", library=str(library)))


@app.post("/query")
def query_route():
    library = _library()
    text = request.form.get("text", "").strip()
    region = request.form.get("region", "深圳").strip() or "深圳"
    item_type = request.form.get("item_type") or None
    if not text:
        return render_template("result.html", title="查询失败", message="请输入查询内容。", back=url_for("index", library=str(library)))
    result = query(library, text, region, item_type)
    grouped_rows = {"labor": [], "material": [], "machinery": [], "equipment": [], "other": []}
    for row in result["top_rows"]:
        key = row.get("原始类型") or "other"
        if key not in grouped_rows:
            key = "other"
        grouped_rows[key].append(row)
    return render_template(
        "query_result.html",
        result=result,
        grouped_rows=grouped_rows,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/prices", methods=["GET", "POST"])
def prices_route():
    library = _library() if request.method == "POST" else _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    filters = {
        "q": request.values.get("q", "").strip(),
        "item_type": request.values.get("item_type", ""),
        "region": request.values.get("region", ""),
        "source_type": request.values.get("source_type", ""),
        "limit": min(int(request.values.get("limit", "200") or 200), 5000),
    }
    rows = []
    source_types = []
    regions = []
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            source_types = [r[0] for r in conn.execute("SELECT DISTINCT source_type FROM prices ORDER BY source_type").fetchall()]
            regions = [r[0] for r in conn.execute("SELECT DISTINCT region FROM prices ORDER BY region").fetchall()]
            sql, params = _price_filter_query(filters, int(filters["limit"]))
            rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    return render_template(
        "prices.html",
        rows=rows,
        filters=filters,
        source_types=source_types,
        regions=regions,
        quality=_price_data_quality(library),
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/export/prices", methods=["GET"])
def export_prices_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    filters = {
        "q": request.args.get("q", "").strip(),
        "item_type": request.args.get("item_type", ""),
        "region": request.args.get("region", ""),
        "source_type": request.args.get("source_type", ""),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "价格清单"
    headers = ["类型", "名称", "规格", "单位", "单价", "地区", "年份/日期", "来源", "口径", "来源文件", "Sheet", "行号", "异常原因"]
    header_fill = PatternFill(start_color="00957F", end_color="00957F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            sql, params = _price_filter_query(filters, 5000)
            for row_index, row in enumerate(conn.execute(sql, params).fetchall(), 2):
                values = [
                    row["item_type"],
                    row["standard_name"] or row["original_name"],
                    row["standard_spec"] or row["original_spec"],
                    row["standard_unit"] or row["original_unit"],
                    row["unit_price"],
                    row["region"],
                    row["price_date"] or row["price_year"],
                    row["source_type"],
                    row["price_scope"] or "",
                    row["file_name"] or "",
                    row["source_sheet"],
                    row["source_row"],
                    row["outlier_reason"] or "",
                ]
                for col_index, value in enumerate(values, 1):
                    ws.cell(row=row_index, column=col_index, value=value)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"价格清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/quality", methods=["GET"])
def export_quality_route():
    library = _resolve_library(request.args.get("library") or str(DEFAULT_LIBRARY))
    quality = _price_data_quality(library)
    wb = Workbook()
    summary = wb.active
    summary.title = "数据质量概览"
    headers = ["校验项", "数量", "处理建议"]
    rows = [
        ["必填项缺失", quality["missing_required"], "补齐名称、单位、地区、类型后再用于正式组价"],
        ["无效单价", quality["invalid_price"], "检查单价为空、0 或负数的记录"],
        ["疑似重复记录", quality["duplicate_items"], "核对同名称、规格、单位、地区、单价的重复来源"],
        ["已标记异常", quality["outliers"], "复核异常原因，确认后再发布使用"],
    ]
    header_fill = PatternFill(start_color="00957F", end_color="00957F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_index, header in enumerate(headers, 1):
        cell = summary.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, 2):
        for col_index, value in enumerate(row, 1):
            summary.cell(row=row_index, column=col_index, value=value)

    detail = wb.create_sheet("问题明细")
    detail_headers = ["问题类型", "类型", "名称", "规格", "单位", "单价", "地区", "来源文件", "Sheet", "行号", "异常原因"]
    for col_index, header in enumerate(detail_headers, 1):
        cell = detail.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    if (library / "database" / "cost_database.sqlite").exists():
        with connect(library) as conn:
            issue_sql = """
                SELECT p.*, f.original_name AS file_name,
                    CASE
                        WHEN p.unit_price IS NULL OR p.unit_price <= 0 THEN '无效单价'
                        WHEN COALESCE(p.original_name, p.standard_name, '') = ''
                          OR COALESCE(p.original_unit, p.standard_unit, '') = ''
                          OR COALESCE(p.region, '') = ''
                          OR COALESCE(p.item_type, '') = '' THEN '必填项缺失'
                        WHEN p.is_outlier = 1 THEN '已标记异常'
                        ELSE '需复核'
                    END AS issue_type
                FROM prices p
                LEFT JOIN files f ON f.id = p.file_id
                WHERE p.unit_price IS NULL OR p.unit_price <= 0
                   OR COALESCE(p.original_name, p.standard_name, '') = ''
                   OR COALESCE(p.original_unit, p.standard_unit, '') = ''
                   OR COALESCE(p.region, '') = ''
                   OR COALESCE(p.item_type, '') = ''
                   OR p.is_outlier = 1
                ORDER BY p.created_at DESC
                LIMIT 5000
            """
            for row_index, row in enumerate(conn.execute(issue_sql).fetchall(), 2):
                values = [
                    row["issue_type"],
                    row["item_type"],
                    row["standard_name"] or row["original_name"] or "",
                    row["standard_spec"] or row["original_spec"] or "",
                    row["standard_unit"] or row["original_unit"] or "",
                    row["unit_price"],
                    row["region"],
                    row["file_name"] or "",
                    row["source_sheet"],
                    row["source_row"],
                    row["outlier_reason"] or "",
                ]
                for col_index, value in enumerate(values, 1):
                    detail.cell(row=row_index, column=col_index, value=value)
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"数据质量报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/price-file")
def price_file_route():
    library = _library()
    init_library(library)
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return render_template("result.html", title="补价失败", message="请上传 Excel 文件。", back=url_for("index", library=str(library)))
    required_fields = {
        "project_name": "项目名称",
        "region": "地区",
        "project_type": "项目类型",
        "price_scope": "价格口径",
    }
    validation_errors = [label for field, label in required_fields.items() if not (request.form.get(field) or "").strip()]
    if validation_errors:
        return render_template(
            "result.html",
            title="导入校验未通过",
            message="请补充必填信息后重新上传。",
            errors=[f"{label}不能为空" for label in validation_errors],
            back=url_for("index", library=str(library)),
        )
    upload_dir = library / "temp" / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    input_path = upload_dir / f"{today_compact()}_{safe_filename(uploaded.filename)}"
    uploaded.save(input_path)
    result = price_file(
        library=library,
        input_path=input_path,
        project_name=request.form.get("project_name") or "新项目",
        region=request.form.get("region") or "深圳",
        project_type=request.form.get("project_type") or "住宅",
        price_scope=request.form.get("price_scope") or "含税",
        payment_terms=request.form.get("payment_terms") or "",
        duration=request.form.get("duration") or "",
    )
    return render_template(
        "price_result.html",
        result=result,
        stats=_stats(library),
        back=url_for("index", library=str(library)),
        task_url=url_for("task_detail_route", task_id=result["task_id"], library=str(library)),
    )


@app.route("/import/shaanxi", methods=["GET", "POST"])
def import_shaanxi_route():
    library = _library()
    init_library(library)
    if request.method == "GET":
        return render_template("shaanxi_import.html", stats=_stats(library), back=url_for("index", library=str(library)))
    source_url = request.form.get("source_url", "").strip()
    pdf_url = request.form.get("pdf_url", "").strip()
    pdf_path = request.form.get("pdf_path", "").strip()
    result = import_shaanxi_price(
        library=library,
        source_url=source_url,
        pdf_url=pdf_url,
        pdf_path=Path(pdf_path) if pdf_path else None,
    )
    return render_template(
        "import_result.html",
        result={"summary": result.summary, "log": result.log},
        stats=_stats(library),
        back=url_for("index", library=str(library)),
    )


@app.route("/download", methods=["GET"])
def download():
    path = Path(request.args["path"])
    return send_file(path, as_attachment=True)


@app.route("/open-folder", methods=["GET"])
def open_folder():
    path = Path(request.args["path"])
    folder = path if path.is_dir() else path.parent
    os.startfile(folder)  # type: ignore[attr-defined]
    return redirect(url_for("index", library=str(DEFAULT_LIBRARY)))


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass
    init_library(DEFAULT_LIBRARY)
    print(f"[web_app] file={__file__}")
    print(f"[web_app] cwd={Path.cwd()}")
    print("[web_app] routes=")
    for rule in app.url_map.iter_rules():
        print(f"[web_app]   {rule}")
    url = "http://127.0.0.1:8765"
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    server = make_server("127.0.0.1", 8765, app, threaded=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
